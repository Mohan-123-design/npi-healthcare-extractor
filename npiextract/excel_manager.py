# excel_manager.py - Excel Output with Full Formatting

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import pandas as pd
import os
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelManager:

    def write_results(self, results: list, filepath: str):
        """Create formatted multi-sheet Excel file"""
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Main Results
        ws = wb.active
        ws.title = "NPI Results"
        self._write_main(ws, results)
        
        # Sheet 2: Summary
        ws2 = wb.create_sheet("Summary")
        self._write_summary(ws2, results)
        
        # Sheet 3: Not Found
        ws3 = wb.create_sheet("Not Found")
        self._write_not_found(ws3, results)
        
        # Sheet 4: Validated NPIs
        ws4 = wb.create_sheet("Validated NPIs")
        self._write_validated(ws4, results)
        
        wb.save(filepath)
        logger.info(f"Excel saved: {filepath}")

    def _thin_border(self):
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def _write_main(self, ws, results):
        # Title
        ws.merge_cells('A1:M1')
        c = ws['A1']
        c.value = "NPI EXTRACTION RESULTS"
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells('A2:M2')
        c2 = ws['A2']
        found = sum(1 for r in results if r.get("npi_found"))
        c2.value = (
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
            f"Total: {len(results)} | Found: {found} | "
            f"Success Rate: {found/len(results)*100:.1f}%" if results else "No data"
        )
        c2.font = Font(italic=True, size=9, color="FFFFFF")
        c2.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        c2.alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "S.No", "Provider URL", "NPI Number", "Extraction Method",
            "Confidence", "Validation", "Registry Name", "Specialty",
            "State", "API Used", "All Candidates", "Error", "Timestamp"
        ]
        
        header_colors = [
            "BDD7EE", "BDD7EE", "C6EFCE", "D9E1F2",
            "FFEB9C", "D9E1F2", "D9E1F2", "D9E1F2",
            "D9E1F2", "D9E1F2", "D9E1F2", "FFC7CE", "D9E1F2"
        ]

        for ci, (h, color) in enumerate(zip(headers, header_colors), 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()
        ws.row_dimensions[3].height = 22

        # Data rows
        for ri, result in enumerate(results, 4):
            candidates_str = " | ".join([
                c["npi"] for c in result.get("all_candidates", [])
            ][:3])
            
            row_data = [
                ri - 3,
                result.get("url", ""),
                result.get("npi_found", "NOT FOUND"),
                result.get("extraction_method", ""),
                f"{result.get('confidence', 0)}%",
                result.get("validation_status", ""),
                result.get("registry_name", ""),
                result.get("registry_specialty", ""),
                result.get("registry_state", ""),
                result.get("api_used", ""),
                candidates_str,
                result.get("error", "")[:100] if result.get("error") else "",
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ]

            has_npi = bool(result.get("npi_found"))
            
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
                cell.border = self._thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=ci == 2)
                
                # Color NPI column
                if ci == 3:
                    if has_npi:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")
                elif ri % 2 == 0 and ci != 3:
                    cell.fill = PatternFill(start_color="F7FBFF", end_color="F7FBFF", fill_type="solid")
            
            ws.row_dimensions[ri].height = 18

        # Column widths
        widths = [6, 55, 14, 18, 10, 14, 28, 22, 8, 12, 30, 25, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:M{ws.max_row}"

    def _write_summary(self, ws, results):
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

        ws.merge_cells('A1:B1')
        ws['A1'] = "EXTRACTION SUMMARY"
        ws['A1'].font = Font(bold=True, size=13, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")

        total = len(results)
        found = sum(1 for r in results if r.get("npi_found"))
        validated = sum(1 for r in results if r.get("validation_status") == "valid")

        method_counts = {}
        api_counts = {}
        for r in results:
            m = r.get("extraction_method") or "none"
            a = r.get("api_used") or "none"
            method_counts[m] = method_counts.get(m, 0) + 1
            api_counts[a] = api_counts.get(a, 0) + 1

        rows = [
            ("Total URLs", total),
            ("NPI Found", found),
            ("NPI Not Found", total - found),
            ("Success Rate", f"{found/total*100:.1f}%" if total else "0%"),
            ("Validated by Registry", validated),
            ("", ""),
            ("Extraction Methods", ""),
        ]
        for m, c in method_counts.items():
            rows.append((f"  {m}", c))
        rows.append(("", ""))
        rows.append(("API Usage", ""))
        for a, c in api_counts.items():
            rows.append((f"  {a}", c))

        for i, (label, val) in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=val)

    def _write_not_found(self, ws, results):
        not_found = [r for r in results if not r.get("npi_found")]
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30

        ws['A1'] = f"URLs Where NPI Not Found ({len(not_found)})"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        ws['A2'] = "URL"
        ws['B2'] = "API Used"
        ws['C2'] = "Error"

        for i, r in enumerate(not_found, 3):
            ws.cell(row=i, column=1, value=r.get("url", ""))
            ws.cell(row=i, column=2, value=r.get("api_used", ""))
            ws.cell(row=i, column=3, value=r.get("error", "")[:100] if r.get("error") else "")

    def _write_validated(self, ws, results):
        validated = [r for r in results if r.get("validation_status") == "valid"]
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10

        ws['A1'] = f"Validated NPIs ({len(validated)})"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")

        for col, header in enumerate(["NPI", "Registry Name", "Specialty", "URL", "State"], 1):
            ws.cell(row=2, column=col, value=header).font = Font(bold=True)

        for i, r in enumerate(validated, 3):
            ws.cell(row=i, column=1, value=r.get("npi_found", ""))
            ws.cell(row=i, column=2, value=r.get("registry_name", ""))
            ws.cell(row=i, column=3, value=r.get("registry_specialty", ""))
            ws.cell(row=i, column=4, value=r.get("url", ""))
            ws.cell(row=i, column=5, value=r.get("registry_state", ""))